"""
servicios/reportes.py — Generación de reportes (Excel y gráficos).

Contiene funciones independientes para generar archivos exportables.
No depende de Telegram, solo de datos y librerías de visualización.
"""
import matplotlib
matplotlib.use('Agg')

import io
import asyncio
import logging
from datetime import datetime

import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

from db import db
from utils.textos import MESES_ES

logger = logging.getLogger('Manguito-Reportes')


def nombre_mes_es() -> str:
    """Retorna el mes actual en español con el año. Ej: 'Febrero 2026'."""
    ahora = datetime.now()
    return f"{MESES_ES[ahora.month]} {ahora.year}"


async def generar_resumen_imagen(user_id: int):
    """
    Genera una imagen PNG tipo 'card' dark mode con el resumen mensual.
    Ideal para compartir en Instagram Stories, redes sociales, etc.
    
    Returns:
        BytesIO con la imagen PNG, o None si no hay datos.
    """
    try:
        ingresos, gastos = await db.get_resumen_mensual(user_id)
        if ingresos == 0 and gastos == 0:
            return None

        saldo = ingresos - gastos
        datos_grafico = await db.get_datos_analisis(user_id)
        mes = nombre_mes_es()

        def _render():
            fig, ax = plt.subplots(figsize=(7.2, 7.2), facecolor='#0f0f23')
            ax.set_facecolor('#0f0f23')
            ax.axis('off')

            # Título
            ax.text(0.5, 0.95, '🥭 MANGUITO', ha='center', va='top',
                    color='#00D4AA', fontsize=24, fontweight='bold',
                    transform=ax.transAxes)
            ax.text(0.5, 0.88, f'Resumen de {mes}', ha='center', va='top',
                    color='#888888', fontsize=13, transform=ax.transAxes)

            # Línea separadora
            ax.plot([0.1, 0.9], [0.84, 0.84], color='#333333', linewidth=1,
                    transform=ax.transAxes)

            # Balance principal
            color_saldo = '#00D4AA' if saldo >= 0 else '#FF6B6B'
            emoji_saldo = '📈' if saldo >= 0 else '📉'
            ax.text(0.5, 0.76, f'{emoji_saldo} Balance: ${saldo:,.0f}',
                    ha='center', va='top', color=color_saldo,
                    fontsize=20, fontweight='bold', transform=ax.transAxes)

            # Ingresos y Gastos
            ax.text(0.25, 0.66, f'Ingresos', ha='center', va='top',
                    color='#888888', fontsize=11, transform=ax.transAxes)
            ax.text(0.25, 0.60, f'${ingresos:,.0f}', ha='center', va='top',
                    color='#34D399', fontsize=16, fontweight='bold',
                    transform=ax.transAxes)

            ax.text(0.75, 0.66, f'Gastos', ha='center', va='top',
                    color='#888888', fontsize=11, transform=ax.transAxes)
            ax.text(0.75, 0.60, f'${gastos:,.0f}', ha='center', va='top',
                    color='#FF6B6B', fontsize=16, fontweight='bold',
                    transform=ax.transAxes)

            # Línea separadora
            ax.plot([0.1, 0.9], [0.54, 0.54], color='#333333', linewidth=1,
                    transform=ax.transAxes)

            # Top categorías
            if datos_grafico:
                ax.text(0.5, 0.49, 'Top Categorias', ha='center', va='top',
                        color='#888888', fontsize=12, transform=ax.transAxes)

                total_gastos = sum(m for _, m in datos_grafico) or 1
                colores_cat = ['#00D4AA', '#4ECDC4', '#FFE66D', '#FF6B6B', '#A78BFA']
                y_pos = 0.42
                for i, (cat, monto) in enumerate(datos_grafico[:5]):
                    pct = (monto / total_gastos) * 100
                    color = colores_cat[i % len(colores_cat)]

                    # Barra de progreso
                    bar_width = min(pct / 100 * 0.5, 0.5)
                    ax.barh(y_pos, bar_width, height=0.035, left=0.1,
                            color=color, alpha=0.8, transform=ax.transAxes)
                    ax.text(0.1, y_pos + 0.02, f'{cat}', ha='left', va='bottom',
                            color='#e0e0e0', fontsize=10, transform=ax.transAxes)
                    ax.text(0.9, y_pos + 0.015, f'${monto:,.0f} ({pct:.0f}%)',
                            ha='right', va='center', color='#999999',
                            fontsize=9, transform=ax.transAxes)
                    y_pos -= 0.065

            # Footer
            ax.text(0.5, 0.04, 'Hecho con Manguito Bot 🥭',
                    ha='center', va='bottom', color='#444444',
                    fontsize=9, style='italic', transform=ax.transAxes)

            plt.tight_layout()
            buf = io.BytesIO()
            plt.savefig(buf, format='png', facecolor='#0f0f23', dpi=150,
                        bbox_inches='tight')
            buf.seek(0)
            plt.close()
            return buf

        return await asyncio.to_thread(_render)

    except Exception as e:
        logger.error(f"Error generando imagen resumen: {e}")
        plt.close()
        return None


def generar_grafico_premium(datos_grafico: list):
    """
    Genera el gráfico de dona dark mode.

    Args:
        datos_grafico: lista de tuplas (categoria, monto_total).

    Returns:
        BytesIO con la imagen PNG, o None si falla.
    """
    try:
        categorias = [f[0] for f in datos_grafico]
        montos = [f[1] for f in datos_grafico]
        total = sum(montos)
        mes = nombre_mes_es()

        colores = [
            '#00D4AA', '#FF6B6B', '#4ECDC4', '#FFE66D', '#A78BFA',
            '#FB923C', '#F472B6', '#34D399', '#60A5FA', '#F87171'
        ]

        fig, ax = plt.subplots(figsize=(7, 7.5), facecolor='#0f0f23')
        ax.set_facecolor('#0f0f23')

        wedges, texts, autotexts = ax.pie(
            montos, labels=None, autopct='%1.0f%%', startangle=140,
            colors=colores[:len(categorias)], pctdistance=0.78,
            wedgeprops=dict(width=0.35, edgecolor='#0f0f23', linewidth=2.5)
        )

        for t in autotexts:
            t.set_color('white')
            t.set_fontsize(10)
            t.set_fontweight('bold')

        # Leyenda SIN emojis (matplotlib no los renderiza bien)
        leyenda_labels = []
        for cat, monto in zip(categorias, montos):
            pct = (monto / total * 100) if total > 0 else 0
            leyenda_labels.append(f"{cat}  ${monto:,.0f} ({pct:.0f}%)")

        legend = ax.legend(
            wedges, leyenda_labels, loc='lower center', ncol=2,
            fontsize=9, frameon=False, bbox_to_anchor=(0.5, -0.08)
        )
        for text in legend.get_texts():
            text.set_color('#e0e0e0')

        ax.set_title('Manguito', color='#00D4AA', fontsize=18, fontweight='bold', pad=25)
        ax.text(0, 1.15, mes, ha='center', va='center', color='#888888',
                fontsize=11, transform=ax.transAxes)

        # Círculo central con total
        centro = plt.Circle((0, 0), 0.32, fc='#0f0f23')
        ax.add_artist(centro)
        ax.text(0, 0.05, f'${total:,.0f}', ha='center', va='center',
                color='#00D4AA', fontsize=16, fontweight='bold')
        ax.text(0, -0.12, 'TOTAL', ha='center', va='center',
                color='#666666', fontsize=9)

        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format='png', facecolor='#0f0f23', dpi=150, bbox_inches='tight')
        buf.seek(0)
        plt.close()
        return buf
    except Exception as e:
        logger.error(f"Error generando gráfico: {e}")
        plt.close()
        return None


async def generar_excel_bytes(user_id: int):
    """
    Genera un archivo Excel (.xlsx) con los movimientos del usuario.

    Incluye:
    - Hoja 1: Movimientos detallados con formato profesional
    - Hoja 2: Dashboard con resumen, tabla por categoría y gráficos

    Returns:
        BytesIO con el archivo Excel, o None si no hay datos.
    """
    datos = await db.get_all_user_data(user_id)

    if not datos:
        return None

    # La construcción del Excel es CPU-bound → lo mandamos a un thread
    return await asyncio.to_thread(_construir_excel, datos)


def _construir_excel(datos):
    """Construye el Excel (CPU-bound, se ejecuta en thread pool)."""

    wb = Workbook()

    # --- Estilos reutilizables ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ===========================
    # HOJA 1: MOVIMIENTOS
    # ===========================
    ws = wb.active
    ws.title = "Movimientos"

    headers = ["ID", "Fecha", "Tipo", "Monto", "Categoría", "Descripción"]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Formatos de fecha soportados
    formatos_fecha = [
        '%Y-%m-%d %H:%M:%S', '%d-%m-%Y %H:%M', '%d/%m/%Y %H:%M',
        '%Y-%m-%d', '%d-%m-%Y',
    ]

    fill_par = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for i, row in enumerate(datos, start=2):
        try:
            r_id = row[0]

            # Convertir fecha string a datetime
            r_fecha = row[2]
            for fmt in formatos_fecha:
                try:
                    r_fecha = datetime.strptime(str(row[2]), fmt)
                    break
                except ValueError:
                    continue

            r_tipo = row[3].capitalize()
            r_monto = row[4]
            r_cat = row[5]
            r_desc = row[6]

            ws.append([r_id, r_fecha, r_tipo, r_monto, r_cat, r_desc])

            # Formato: ID
            ws.cell(row=i, column=1).alignment = center_align
            ws.cell(row=i, column=1).border = border

            # Formato: Fecha
            cell_fecha = ws.cell(row=i, column=2)
            cell_fecha.number_format = 'dd/mm/yyyy hh:mm'
            cell_fecha.alignment = center_align
            cell_fecha.border = border

            # Formato: Tipo
            ws.cell(row=i, column=3).alignment = center_align
            ws.cell(row=i, column=3).border = border

            # Formato: Monto
            cell_monto = ws.cell(row=i, column=4)
            cell_monto.number_format = '"$"#,##0'
            cell_monto.alignment = Alignment(horizontal="right")
            cell_monto.border = border

            if r_tipo == "Ingreso":
                cell_monto.font = Font(color="006100", bold=False)
            else:
                cell_monto.font = Font(color="9C0006", bold=False)

            # Formato: Categoría
            ws.cell(row=i, column=5).border = border

            # Formato: Descripción
            cell_desc = ws.cell(row=i, column=6)
            cell_desc.alignment = Alignment(wrap_text=True)
            cell_desc.border = border

            # Zebra striping
            if i % 2 == 0:
                for col_idx in range(1, 7):
                    ws.cell(row=i, column=col_idx).fill = fill_par

        except Exception:
            continue

    # Ocultar columna ID
    ws.column_dimensions['A'].hidden = True

    # Anchos de columna
    anchos = {'B': 18, 'C': 10, 'D': 15, 'E': 20, 'F': 80}
    for col_letter, width in anchos.items():
        ws.column_dimensions[col_letter].width = width

    # Filtros
    ws.auto_filter.ref = ws.dimensions

    # ===========================
    # HOJA 2: DASHBOARD
    # ===========================
    ws_dash = wb.create_sheet("Resumen y Gráficos")

    # Calcular totales
    gastos_por_cat = {}
    ingresos_total = 0
    egresos_total = 0

    for row in datos:
        monto = row[4]
        if row[3].lower() == 'egreso':
            cat = row[5]
            gastos_por_cat[cat] = gastos_por_cat.get(cat, 0) + monto
            egresos_total += monto
        elif row[3].lower() == 'ingreso':
            ingresos_total += monto

    # Tabla: Resumen General
    ws_dash.append(["RESUMEN GENERAL", ""])
    ws_dash.append(["Total Ingresos", ingresos_total])
    ws_dash.append(["Total Gastos", egresos_total])
    ws_dash.append(["Balance", ingresos_total - egresos_total])
    ws_dash.append(["", ""])

    # Estilos Resumen General
    ws_dash['A1'].font = Font(bold=True, size=12)
    ws_dash['B2'].font = Font(color="006100", bold=True)
    ws_dash['B3'].font = Font(color="9C0006", bold=True)
    ws_dash['B4'].font = Font(bold=True)
    for r in range(2, 5):
        ws_dash.cell(row=r, column=2).number_format = '"$"#,##0'

    # Tabla: Categorías
    ws_dash.append(["GASTOS POR CATEGORÍA", "Monto"])
    ws_dash.merge_cells('A6:A6')

    ws_dash['A6'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_dash['A6'].fill = header_fill
    ws_dash['B6'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_dash['B6'].fill = header_fill

    for r in range(1, 5):
        for c in range(1, 3):
            ws_dash.cell(row=r, column=c).border = border

    row_idx = 7
    for cat, monto in sorted(gastos_por_cat.items(), key=lambda x: x[1], reverse=True):
        ws_dash.append([cat, monto])
        cell_cat = ws_dash.cell(row=row_idx, column=1)
        cell_monto = ws_dash.cell(row=row_idx, column=2)
        cell_monto.number_format = '"$"#,##0'
        cell_cat.border = border
        cell_monto.border = border
        row_idx += 1

    # --- Datos para gráfico de torta (agrupando menores al 5%) ---
    grafico_data = {}
    otros_monto = 0
    total_gastos_grafico = sum(gastos_por_cat.values())

    if total_gastos_grafico > 0:
        for cat, monto in gastos_por_cat.items():
            pct = monto / total_gastos_grafico
            if pct < 0.05:
                otros_monto += monto
            else:
                grafico_data[cat] = monto

        if otros_monto > 0:
            grafico_data["Otros (Menores)"] = otros_monto

    # Escribir datos en columnas lejanas (AA y AB)
    col_cat_idx = 27  # AA
    col_val_idx = 28  # AB

    ws_dash.cell(row=1, column=col_cat_idx, value="Cat Grafico")
    ws_dash.cell(row=1, column=col_val_idx, value="Monto Grafico")

    g_row = 2
    for cat, monto in sorted(grafico_data.items(), key=lambda x: x[1], reverse=True):
        ws_dash.cell(row=g_row, column=col_cat_idx, value=cat)
        ws_dash.cell(row=g_row, column=col_val_idx, value=monto)
        g_row += 1

    # Crear gráfico de torta
    if grafico_data:
        pie = PieChart()
        pie.title = "Distribución de Gastos"

        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = False
        pie.dataLabels.showCatName = False
        pie.dataLabels.showSerName = False
        pie.dataLabels.showLeaderLines = True
        pie.dataLabels.position = "outEnd"

        count = len(grafico_data)
        labels = Reference(ws_dash, min_col=col_cat_idx, min_row=2, max_row=2 + count - 1)
        data = Reference(ws_dash, min_col=col_val_idx, min_row=2, max_row=2 + count - 1)

        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        ws_dash.add_chart(pie, "D2")
    else:
        ws_dash.cell(row=7, column=4, value="Sin gastos registrados para graficar")

    # --- Gráfico de evolución temporal ---
    gastos_por_dia = {}
    for row in datos:
        if row[3].lower() == 'egreso':
            try:
                fecha_str = str(row[2])
                fecha_obj = None
                for fmt in formatos_fecha:
                    try:
                        fecha_obj = datetime.strptime(fecha_str, fmt)
                        break
                    except ValueError:
                        continue
                if not fecha_obj:
                    continue
                dia = fecha_obj.day
                gastos_por_dia[dia] = gastos_por_dia.get(dia, 0) + row[4]
            except Exception:
                continue

    if gastos_por_dia:
        col_dia_idx = 29  # AC
        col_acum_idx = 30  # AD

        ws_dash.cell(row=1, column=col_dia_idx, value="Día")
        ws_dash.cell(row=1, column=col_acum_idx, value="Gasto Acumulado")

        dias_ordenados = sorted(gastos_por_dia.keys())
        acumulado = 0
        ev_row = 2
        for dia in dias_ordenados:
            acumulado += gastos_por_dia[dia]
            ws_dash.cell(row=ev_row, column=col_dia_idx, value=f"Día {dia}")
            ws_dash.cell(row=ev_row, column=col_acum_idx, value=acumulado)
            ev_row += 1

        # Gráfico de líneas
        line_chart = LineChart()
        line_chart.title = "Evolución de Gastos del Mes"
        line_chart.x_axis.title = "Día del Mes"
        line_chart.y_axis.title = "Gasto Acumulado ($)"
        line_chart.style = 10
        line_chart.y_axis.numFmt = '"$"#,##0'
        line_chart.width = 20
        line_chart.height = 12

        count_dias = len(dias_ordenados)
        data_line = Reference(ws_dash, min_col=col_acum_idx, min_row=1, max_row=1 + count_dias)
        cats_line = Reference(ws_dash, min_col=col_dia_idx, min_row=2, max_row=1 + count_dias)

        line_chart.add_data(data_line, titles_from_data=True)
        line_chart.set_categories(cats_line)

        serie = line_chart.series[0]
        serie.graphicalProperties.line.width = 25000

        ws_dash.add_chart(line_chart, "D18")

    ws_dash.column_dimensions['A'].width = 25
    ws_dash.column_dimensions['B'].width = 15

    # Guardar en buffer
    ex_buffer = io.BytesIO()
    wb.save(ex_buffer)
    ex_buffer.seek(0)
    return ex_buffer
